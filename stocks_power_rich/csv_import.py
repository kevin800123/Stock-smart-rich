"""解析使用者每日上傳的籌碼 CSV → 標準化欄位 → 寫入 chip_snapshot。

自動偵測編碼（cp950/big5hkscs/utf-8/utf-16）並自動定位標頭與資料日期列，
容忍前置列數不同、末欄未加引號逗號、西元或民國年。
"""
import csv
import io
import json
import os
import re
import shutil
from datetime import datetime

# 嘗試的編碼（cp950＝Windows Big5 超集，含「碁」等字；再退 UTF-8/UTF-16）
ENCODINGS = ["cp950", "big5hkscs", "utf-8-sig", "utf-16", "utf-8"]

COLMAP = {
    "代碼": "code", "商品": "name", "成交": "close", "漲幅%": "change_pct",
    "總量": "volume", "市值(億)": "market_cap", "股本(億)": "capital",
    "推估獲利": "est_profit", "蘭質": "lan_score", "LPE": "lpe", "蘭值": "lan_value",
    "W55": "w55", "集保": "custody",
    "大戶增比": "big_holder_ratio", "人數降比": "holder_drop_ratio",
    "月增": "month_inc", "年增": "rev_yoy", "累增": "accum_inc",
    "投三": "trust_3d", "外三": "foreign_3d", "產業": "industry",
    "細產業": "sub_industry", "所有細產業": "all_sub_industry",
    "產業地位": "industry_position",
}
NUMERIC = {
    "close", "change_pct", "volume", "market_cap", "capital", "est_profit",
    "lan_score", "lpe", "lan_value", "w55", "custody", "big_holder_ratio",
    "holder_drop_ratio", "month_inc", "rev_yoy", "accum_inc", "trust_3d", "foreign_3d",
}


def _to_float(v):
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _normalize_field_count(row: list, n: int) -> list:
    """調整欄數至 n：欄數過多時，把溢出欄位併回最後一欄（未加引號的「產業地位」
    可能含 ASCII 逗號而被拆開）；欄數不足時補空字串。"""
    if len(row) > n:
        return row[: n - 1] + [",".join(row[n - 1:])]
    if len(row) < n:
        return row + [""] * (n - len(row))
    return row


def _decode(raw: bytes) -> str:
    """自動偵測編碼：優先選能解出關鍵欄名（代碼、商品）的編碼。"""
    fallback = None
    for enc in ENCODINGS:
        try:
            txt = raw.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
        if "代碼" in txt and "商品" in txt:
            return txt
        if fallback is None:
            fallback = txt
    return fallback if fallback is not None else raw.decode("cp950", errors="replace")


def _cell_str(c) -> str:
    if c is None:
        return ""
    if isinstance(c, float) and c.is_integer():  # 30000.0 → "30000"
        return str(int(c))
    return str(c)


def _read_excel_rows(raw: bytes) -> list:
    """讀 .xlsx/.xlsm 第一個工作表 → 每列為字串 list。"""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = [[_cell_str(c) for c in r] for r in ws.iter_rows(values_only=True)]
    wb.close()
    return out


def _read_rows(path: str) -> list:
    """讀檔成「每列為儲存格 list」，自動分辨 Excel(zip) 與 CSV(多編碼)。"""
    raw = open(path, "rb").read()
    if raw[:4] == b"PK\x03\x04":  # xlsx/xlsm 為 zip
        return _read_excel_rows(raw)
    return list(csv.reader(io.StringIO(_decode(raw))))


def _trim_trailing_empty(cells: list) -> list:
    out = list(cells)
    while out and str(out[-1]).strip() == "":
        out.pop()
    return out


def _find_header_index(rows: list) -> int:
    """掃出標頭列（含「代碼」與「商品」），找不到則視為第一列。"""
    for i, r in enumerate(rows):
        joined = ",".join(_cell_str(c) for c in r)
        if "代碼" in joined and "商品" in joined:
            return i
    return 0


def _extract_date(rows: list) -> str:
    """掃描列找資料日期，支援西元與民國年。"""
    for r in rows:
        joined = " ".join(_cell_str(c) for c in r)
        m = re.search(r"(\d{2,4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", joined)
        if m:
            year = int(m.group(1))
            if year < 1911:  # 民國 → 西元
                year += 1911
            return f"{year:04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return datetime.now().strftime("%Y-%m-%d")


def parse_csv(path: str):
    rows_in = _read_rows(path)
    if not rows_in:
        return datetime.now().strftime("%Y-%m-%d"), []
    hidx = _find_header_index(rows_in)
    snap_date = _extract_date(rows_in[:hidx] if hidx else rows_in[:3])

    header = _trim_trailing_empty([_cell_str(c).strip().strip('"') for c in rows_in[hidx]])
    n = len(header)

    rows = []
    for raw_row in rows_in[hidx + 1:]:
        cells = _trim_trailing_empty([_cell_str(c) for c in raw_row])
        if not cells:
            continue
        rec = dict(zip(header, _normalize_field_count(cells, n)))
        d = {}
        for src, dst in COLMAP.items():
            if src in rec:
                val = rec[src]
                d[dst] = _to_float(val) if dst in NUMERIC else (val.strip() if val else None)
        if not d.get("code"):
            continue
        d["raw_json"] = json.dumps(rec, ensure_ascii=False)
        rows.append(d)
    return snap_date, rows


def find_latest_file(directory: str):
    """回傳資料夾內最新（mtime）的 .csv/.xlsx/.xlsm 檔，無則 None。"""
    import glob

    files = []
    for ext in ("*.csv", "*.xlsx", "*.xlsm"):
        files.extend(glob.glob(os.path.join(directory, ext)))
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def import_csv(conn, path: str, store_dir: str = "data/csv"):
    from .db import insert_chip_snapshot

    snap_date, rows = parse_csv(path)
    insert_chip_snapshot(conn, snap_date, rows)
    os.makedirs(store_dir, exist_ok=True)
    stored = os.path.join(store_dir, f"{snap_date}.csv")
    shutil.copyfile(path, stored)
    conn.execute(
        "INSERT INTO csv_files (snap_date, stored_path, imported_at) VALUES (?,?,?)",
        (snap_date, stored, datetime.now().isoformat()),
    )
    conn.commit()
    return snap_date, len(rows)
